from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.borders import Side
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font =Font(color= "FF0000", italic= True, bold= True) # 글자색은 빨갛게, 이탤릭, 글자굵게
b1.font =Font(color="CC33FF", name= "Arial", strike= True) # 폰트를 Arial로 설정, 색을 보라색으로, 취소선 적용
c1.font =Font(color="0000FF", size=20, underline= "single") # 글자색 파랗게, 사이즈를 20, 밑줄 한줄로 두줄일땐 더블(double)

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")) # 해당 셀에 상하좌우에 테두리를 설정하는 객체 설정
a1.border = thin_border # 해당셀에 적용
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows: # 엑셀을 한줄씩
    for cell in row: # 한칸씩
        cell.alignment = Alignment(horizontal="center", vertical="center") # 각 셀에 대해 중앙정렬 참고로 ""안에는 center, left, right, top, bottom 들갈수 있음


        if cell.column == 1: # A 번호열은 제외
            continue

        # cell 이 정수형 데이터고(영어,수학등은 비교할 필요 없으므로) 90점 보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로 설정
            cell.font = Font(color="FF0000") # 폰트 색상 변경
            
# 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정


wb.save("sample_style.xlsx")
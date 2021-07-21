from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2부터 D2 까지 합치겠음
ws["B2"].value = "MERGEDCELLMERGEDCELL"

wb.save("sample_merge.xlsx")

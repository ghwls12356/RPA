# 셀을 범위로 가져오기
from openpyxl import Workbook
from random import * 

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"]) # append에는  리스트와 튜플 사용가능
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 엑셀에서 영어 점수(column)만  가지고 오기
# print(col_B) # B열의 정보 출력 
for cell in col_B: # col_B 안에서 cell 이 돎
    print(cell.value) # 각각의 영어 점수 출력

col_range =  ws["B:C"] # 영어, 수학 column 함께 가지고 오가
for cols in col_range: # B(영어), C(수학)을 가져옴
    for cell in cols: # 각각의 열 리스트로 가져와서 값을 받아옴
        print(cell.value) # 출력

row_title = ws[1] # 1번째 행(row)만 가지고 오기
for cell in row_title: 
    print(cell.value)

row_range = ws[2:6] # 1번째 줄인 title을 제외하고 2번째 줄에서 6번째 줄까지 가지고 오기
for rows in row_range: # row_range에서 한줄씩 튜플형태로 가져옴
    for cell in rows:  # rows 의 한줄을 값 하나하나 씩 가져옴
        print(cell.value, end=" ") # 출력
    print() # 줄 바꿈

from openpyxl.utils.cell import coordinate_from_string  # 셀 정보 확인

row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range: # 한줄씩
    for cell in rows: # 한 줄안에서 값 하나씩
        # print(cell.value, end=" ") #출력
        
        print(cell.coordinate, end=" ") # 셀 정보 확인 , AZ250 같이 커지면 슬라이싱  힘듬
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end= " ")                          # -> 셀정보 확인, 슬라이싱 용이
        #print(xy[0], end="") # A
        #print(xy[1], end=" ") # 1 == print(cell.coordinate, end=" ") , 슬라이싱 예시
    print()



wb.save("sample.xlsx")

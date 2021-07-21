from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# for row in ws.values:
#     for cell in row:
#         print(cell) # 표시된 값이 아니라 셀에 입력된 그 자체를 가져옴(수식 그대로 가져오고 있음)

# 수식이 아닌 실제 데이터 가지고 옴
# evaluate(계산) 되지 않은 상태의 데이터는 None 이라고 표시 -> 저장하고 계산된 내용을 저장하면 표시됨 -> openpyxl에서 작업한 파일은 최종파일이어야 수식이 계산되지 않아 꼬이는 일이 안생김
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)



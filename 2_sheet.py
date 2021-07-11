from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 추가 생성
ws.title = "Newsheet" # 추가된 새로운 sheet 이름변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet("Givensheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("Crosssheet", 2) # 2 번째  index에 sheet 생성

new_ws = wb["Newsheet"] # Dictionary 형태로 sheet 에 접근

print(wb.sheetnames) # 모든 sheet 이름 확인

# sheet 복사
new_ws["A1"] = "Testing" # A1셀에 testing 입력
target = wb.copy_worksheet(new_ws) # new_ws 객체 시트에 내용복사
target.title = "CopiedSheet" # 복사된 시트의 이름설정 후 맨 뒤에 추가배치




wb.save("sample.xlsx")

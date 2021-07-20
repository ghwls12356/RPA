from openpyxl import load_workbook # 원래 있던 워크북 불러옴
wb = load_workbook("sample.xlsx")  # sample.xlsx 를 워크북으로 불러옴
ws = wb.active # 현재 활성화된 워크시트 사용

for row in ws.iter_rows(min_row= 2): # 2번째 줄 부터 학생들의 성적 불러옴
    # 번호, 영어, 수학
    if int(row[1].value) > 80: # 영어 성적 값이  80점을 초과하면
        print(row[0].value, "번 학생은 영어잘함") #  그 번호 학생 출력

# 변경 후 저장
for row in ws.iter_rows(max_row=1): # 첫번째 과목명들만 불러옴
    for cell in row: # 그 과목명을 하나씩 뜯음
        if cell.value == "영어": # 그과목이 영어면
            cell.value = "컴퓨터" # 컴퓨터로 바꿈
 
wb.save("sample_modified.xlsx") # 바꾼 내용을 "sample_modified.xlsx"로 저장
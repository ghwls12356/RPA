# 이미 존재하는 파일을 읽어와서 작업
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx") # sample.xlsx 파일에서 wb을 불러옴
ws = wb.active # 활성화된 sheet

# cell 데이터 불러오기
for x in range(11, 21):
    for y in range(11, 21):
        print(ws.cell(row=y, column=x).value, end=" ") # 1 2 3 4 ... 한칸씩 띄어주기 위해 "end" 사용
    print() # 한줄씩 띄우기 위해 사용

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=y, column=x).value, end=" ") # 1 2 3 4 ... 한칸씩 띄어주기 위해 "end" 사용
    print() # 한줄씩 띄우기 위해 사용
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 현재까지 작성된 최종 성적 데이터를 넣기
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트")) # 리스트나 튜플형태로 삽입

scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]                     # 점수는 이차원 데이터이므로 리스트안에 튜플 형태로 넣는 방식 채택

for s in scores: 
    ws.append(s)        # 기존 성적 리스트에서 튜플 한 개 씩을 받아와 입력

# 1 퀴즈2 점수를 10점으로 수정
for idx, cell in enumerate(ws["D"]): # enumerate 은 반복문 사용 시 몇 번째 반복문인지 확인이 필요할 수 있습니다. 이때 사용합니다.
                                     # 인덱스 번호와 컬렉션의 원소를 tuple형태로 반환합니다.
    if idx == 0:    # 제목인 경우 skip
        continue  
    cell.value = 10


# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가


ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2): # enumerate 인덱스 시작을 2부터 함
    sum_val = sum(score[1:]) - score[3] + 10 # 총점 데이터 (성적 을 넣기위해 사용) , 퀴즈 2가 10점 된 것을 반영
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
    # SUM(B2:G2)
    # SUM(B3:G3)...
    

    # - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
    grade = None # 성적
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
    
    # 3. 출석이 5 미만인 학생은 총점 상관없이 F
    if score[1] < 5:
        grade = "F"
    
    ws.cell(row=idx, column=9).value = grade # I열에 성적 정보 입력

wb.save("quiz_answer.xlsx")
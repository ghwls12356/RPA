from openpyxl import load_workbook
from random import *
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학

ws.move_range("B1:C11", rows=0, cols=1) # 이동하려는 범위 지정후,  row와 colmn 기준으로 몇 칸이동하는지 입력
ws["B1"].value = "국어" # B1셀에 "국어" 입력
for i in range(1, 11): # 국어 성적에 랜덤값 넣어주기
    ws.append([randint(0, 100)])  # 생성
ws.move_range("A12:A22", rows=-10, cols=1 ) #이동
wb.save("sample_korean.xlsx") # 저장
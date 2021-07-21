import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today() # 오늘 날짜 정보
ws["a2"] = "=sum(1, 2, 3)" # 1 + 2 + 3 = 6(합계)
ws["a3"] = "=average(1, 2, 3)" # 2(평균)

ws["a4"] = 10
ws["a5"] = 20
ws["a6"] = "=sum(a4:a5)" # a4 + a5 셀 30

wb.save("sample_formula.xlsx")

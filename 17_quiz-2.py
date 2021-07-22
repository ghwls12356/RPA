from openpyxl import load_workbook
wb = load_workbook("성적목록.xlsx", data_only=True)
ws = wb.active

ws["H1"] = "총점"
ws["I1"] = "성적"
for col in ws.iter_cols(min_row=2, max_row= 11, min_col=4, max_col=4):
    for cell in col:
        cell.value = 10

for colu in ws.iter_cols(min_row=2, max_row= 11, min_col=8, max_col=8):
    for tg in col:
        for i in range(2, 12):
            Bee = i
            ws["H%d" % Bee] = "=sum(B%d:G%d)" %(Bee, Bee)

for column in ws.iter_cols(min_row=2, max_row= 11, min_col=9, max_col=9):
    for grd in column:
        for g in range(2, 12):
            if ws["B%d" % g].value < 5:
                ws["I%d" % g] = "F"
            elif ws["B%d" % g].value + ws["C%d" % g].value + ws["D%d" % g].value + ws["E%d" % g].value  + ws["F%d" % g].value + ws["G%d" % g].value >= 90:
                ws["I%d" % g] = "A"
            elif ws["B%d" % g].value + ws["C%d" % g].value + ws["D%d" % g].value + ws["E%d" % g].value  + ws["F%d" % g].value + ws["G%d" % g].value >= 80:
                ws["I%d" % g] = "B"    
            elif ws["B%d" % g].value + ws["C%d" % g].value + ws["D%d" % g].value + ws["E%d" % g].value  + ws["F%d" % g].value + ws["G%d" % g].value >= 70:
                ws["I%d" % g] = "C" 
            else:
                ws["I%d" % g] = "D" 

wb.save("scores.xlsx")

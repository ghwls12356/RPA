from openpyxl import Workbook
wb = Workbook() # 새로운 파일 생성
ws = wb.active # 현재 활성화된 시트 가져옴
ws.title = "jinsheet" # 활성화된 시트 이름 변경

# []안의 지정셀에 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6
ws["B4"] = "7"

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 '값'을 출력
print(ws["A10"].value) # 값이 없을 땐 'None'을 출력

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3), ...
print(ws.cell(row=1, column=1).value) # ws["A1"].value 와 같음, 출력
print(ws.cell(column=2, row=1).value) # ws["B1"].value 와 같음

ws.cell(column=3, row=1, value=10) # ws["C1"].value = 10 와 같음
print(ws.cell(column=3, row=1).value) # ws["C1"]의 '값' 출력
c = ws.cell(column=3, row=2, value=20) 
print(c.value) # ws["C2"]의 '값' 출력

from random import * # random 패키지의 모든(*로 표현) 구성요소를 추출(가져옴)

# 반복문을 이용해서 랜덤 숫자 채우기
for x in range(1, 11): # 10개 row
    for y in range(1, 11): # 10개 column
        ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이의 숫자
# 들어가는 순서 
index = 1
for j in range(11, 21): # 10개 row
    for k in range(11, 21): # 10개 column
        ws.cell(row=k, column=j, value=index) # 행과 열 좌표 형태로 순서를 가짐
        index += 1

wb.save("sample.xlsx") # sample.xlsx 파일로 워크북 저장
